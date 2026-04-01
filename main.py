import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import uuid
import datetime
import time
import threading
import webbrowser
import logging
from address_parser import parse_address_string, split_addresses
from ups_api import UPSApiClient
import openpyxl
from openpyxl.styles import Font

UPS_ERROR_MAP = {
    "9510113": "The Ready Time is after the local cutoff for this region. Please select an earlier time (e.g., 09:00 - 11:00 AM).",
    "9510118": "Pickup service is not available for this address/day. Check if it is a holiday or remote area.",
    "9500505": "Invalid State/Province code. Even with normalization, UPS requires a 2-letter ISO code.",
    "120120": "Account mismatch. A Canadian account cannot generate domestic US return labels. Falling back to default tracking.",
    "9500781": "The pickup has zero balance for charges (Success)."
}

class UPSPickupGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("UPS Pickup Automation")
        self.root.geometry("800x800")
        
        self.api_client = UPSApiClient()
        self.stop_batch = False
        self.history_file = "pickup_history.json"
        self.manual_mode = tk.BooleanVar(value=True)
        self.next_step_event = threading.Event()
        
        # Trace to auto-resume when unchecking Manual Mode
        self.manual_mode.trace_add("write", lambda *args: self.next_step_event.set() if not self.manual_mode.get() else None)
        
        self.setup_ui()

    def log_message(self, message):
        """Simplistic logger for the GUI."""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {message}")

    def setup_ui(self):
        # macOS-like Aesthetic Tokens
        bg_main = "#F5F5F7"
        bg_card = "#FFFFFF"
        accent_blue = "#007AFF"
        text_primary = "#1D1D1F"
        text_secondary = "#86868B"
        
        # Determine best available font
        available_fonts = ["San Francisco", "Helvetica Neue", "Helvetica", "Segoe UI", "Arial"]
        main_font_family = "Helvetica" # Default
        for f in available_fonts:
            # Simple check by creating a dummy font
            try:
                tk.font.Font(family=f)
                main_font_family = f
                break
            except: continue

        self.main_font = (main_font_family, 10)
        self.bold_font = (main_font_family, 10, "bold")
        self.header_font = (main_font_family, 13, "bold")
        
        style = ttk.Style()
        style.theme_use('clam')
        
        self.root.configure(bg=bg_main)
        
        # Configure Styles
        style.configure("TFrame", background=bg_main)
        style.configure("Card.TFrame", background=bg_card, relief=tk.FLAT)
        
        style.configure("TLabel", background=bg_main, foreground=text_primary, font=self.main_font)
        style.configure("Secondary.TLabel", background=bg_main, foreground=text_secondary, font=(f"{main_font_family}", 9))
        style.configure("Header.TLabel", background=bg_main, foreground=text_primary, font=self.header_font)
        style.configure("Card.TLabel", background=bg_card, foreground=text_primary, font=self.main_font)
        style.configure("CardHeader.TLabel", background=bg_card, foreground=text_primary, font=self.bold_font)
        
        style.configure("TButton", font=self.bold_font, padding=[12, 6], background=bg_card, foreground=text_primary)
        style.map("TButton", background=[("active", "#E8E8ED")])
        
        style.configure("Accent.TButton", font=self.bold_font, padding=[12, 6], background=accent_blue, foreground="white")
        style.map("Accent.TButton", background=[("active", "#0051D7")])
        
        style.configure("TNotebook", background=bg_main, borderwidth=0)
        style.configure("TNotebook.Tab", background=bg_main, foreground=text_secondary, font=self.bold_font, padding=[15, 8])
        style.map("TNotebook.Tab", 
                  background=[("selected", bg_main)], 
                  foreground=[("selected", accent_blue)])
        
        style.configure("Treeview", font=self.main_font, rowheight=28, background=bg_card, fieldbackground=bg_card, borderwidth=0)
        style.configure("Treeview.Heading", font=self.bold_font, background="#E8E8ED", borderwidth=0)
        
        # Main Layout
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=(10, 20))
        
        self.single_tab = ttk.Frame(self.notebook)
        self.batch_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.single_tab, text="Single Pickup")
        self.notebook.add(self.batch_tab, text="Batch Processing")
        
        self.setup_single_tab()
        self.setup_batch_tab()
        
        # Modern Status Bar (Minimal)
        self.status_var = tk.StringVar(value="Ready")
        status_frame = ttk.Frame(self.root, padding=(20, 5))
        status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        status_dot = tk.Label(status_frame, text="●", fg="#34C759", bg=bg_main, font=("Segoe UI", 8))
        status_dot.pack(side=tk.LEFT)
        status_bar = ttk.Label(status_frame, textvariable=self.status_var, font=(f"{main_font_family}", 9))
        status_bar.pack(side=tk.LEFT, padx=5)

    def setup_single_tab(self):
        main_frame = ttk.Frame(self.single_tab, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Address Paste Area
        ttk.Label(main_frame, text="Paste Address:", font=self.header_font).pack(anchor=tk.W, pady=(0, 5))
        self.address_text = tk.Text(main_frame, height=4, font=self.main_font, borderwidth=1, relief=tk.SOLID)
        self.address_text.pack(fill=tk.X, pady=(0, 15))
        
        parse_btn = ttk.Button(main_frame, text="Parse & Audit", command=self.parse_and_fill)
        parse_btn.pack(anchor=tk.W, pady=(0, 20))
        
        # Auditable Fields (Card Style)
        fields_card = ttk.Frame(main_frame, style="Card.TFrame", padding=20)
        fields_card.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(fields_card, text="Audit Details", style="CardHeader.TLabel").grid(row=0, column=0, columnspan=4, sticky=tk.W, pady=(0, 15))
        
        self.fields = {}
        field_labels = ["Street", "Suite", "City", "State", "Zip", "Country", "CompanyName", "ContactName", "Phone", "Email"]
        
        for i, label in enumerate(field_labels):
            row = (i // 2) + 1
            col = (i % 2) * 2
            ttk.Label(fields_card, text=f"{label}:", style="Card.TLabel").grid(row=row, column=col, sticky=tk.W, pady=8, padx=(0, 5))
            entry = ttk.Entry(fields_card, width=28)
            entry.grid(row=row, column=col+1, sticky=tk.W, padx=(0, 20), pady=8)
            self.fields[label] = entry
            
        # Default values
        self.fields["Country"].insert(0, "CA")
        self.fields["CompanyName"].insert(0, "Omnitrans Inc")
        self.fields["ContactName"].insert(0, "Warehouse")
            
        # Date / Time Fields
        datetime_frame = ttk.Frame(main_frame, padding=(0, 20))
        datetime_frame.pack(fill=tk.X)
        
        ttk.Separator(datetime_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(0, 20))
        
        dt_inner = ttk.Frame(datetime_frame)
        dt_inner.pack(fill=tk.X)
        
        ttk.Label(dt_inner, text="Date:").pack(side=tk.LEFT)
        today = datetime.datetime.now()
        dates = [(today + datetime.timedelta(days=i)).strftime("%Y-%m-%d") for i in range(14)]
        self.date_cb = ttk.Combobox(dt_inner, values=dates, state="readonly", width=12)
        self.date_cb.set(dates[0])
        self.date_cb.pack(side=tk.LEFT, padx=(5, 20))
        
        ttk.Label(dt_inner, text="Ready:").pack(side=tk.LEFT)
        times = [f"{str(h).zfill(2)}:00" for h in range(8, 19)]
        self.ready_cb = ttk.Combobox(dt_inner, values=times, state="readonly", width=8)
        self.ready_cb.set("09:00")
        self.ready_cb.pack(side=tk.LEFT, padx=(5, 20))
        
        ttk.Label(dt_inner, text="Close:").pack(side=tk.LEFT)
        self.close_cb = ttk.Combobox(dt_inner, values=times, state="readonly", width=8)
        self.close_cb.set("17:00")
        self.close_cb.pack(side=tk.LEFT, padx=(5, 0))
        
        # Action Buttons
        btn_frame = ttk.Frame(main_frame, padding=(0, 20))
        btn_frame.pack(fill=tk.X)
        
        submit_btn = ttk.Button(btn_frame, text="Generate Pickup", style="Accent.TButton", command=self.submit_pickup)
        submit_btn.pack(side=tk.RIGHT)
        
        cancel_btn = ttk.Button(btn_frame, text="Cancel Pickup", command=self.cancel_pickup_gui)
        cancel_btn.pack(side=tk.RIGHT, padx=10)
        
        clear_btn = ttk.Button(btn_frame, text="Clear", command=self.clear_fields)
        clear_btn.pack(side=tk.LEFT)
        
        history_btn = ttk.Button(btn_frame, text="View History", command=self.show_history_window)
        history_btn.pack(side=tk.LEFT, padx=10)

    def setup_batch_tab(self):
        main_frame = ttk.Frame(self.batch_tab, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Address Paste Area
        ttk.Label(main_frame, text="Batch Addresses:", font=self.header_font).pack(anchor=tk.W, pady=(0, 5))
        self.batch_text = tk.Text(main_frame, height=5, font=self.main_font, borderwidth=1, relief=tk.SOLID)
        self.batch_text.pack(fill=tk.X, pady=(0, 15))

        # Batch Defaults (Card Style)
        defaults_card = ttk.Frame(main_frame, style="Card.TFrame", padding=20)
        defaults_card.pack(fill=tk.X, pady=10)
        
        ttk.Label(defaults_card, text="Processing Defaults", style="CardHeader.TLabel").grid(row=0, column=0, columnspan=6, sticky=tk.W, pady=(0, 15))
        
        self.batch_defaults = {}
        batch_field_labels = ["CompanyName", "ContactName", "Phone", "Email", "ServiceCode"]
        for i, label in enumerate(batch_field_labels):
            row = (i // 3) + 1
            col = (i % 3) * 2
            ttk.Label(defaults_card, text=f"{label}:", style="Card.TLabel").grid(row=row, column=col, sticky=tk.W, pady=8, padx=(0, 5))
            entry = ttk.Entry(defaults_card, width=18)
            entry.grid(row=row, column=col+1, sticky=tk.W, padx=(0, 15), pady=8)
            self.batch_defaults[label] = entry

        # Defaults set
        self.batch_defaults["CompanyName"].insert(0, "Omnitrans Inc")
        self.batch_defaults["ContactName"].insert(0, "Warehouse")
        self.batch_defaults["Phone"].insert(0, "5142886664")
        self.batch_defaults["ServiceCode"].insert(0, "011")

        # Batch Date/Time / Country
        time_country_frame = ttk.Frame(main_frame)
        time_country_frame.pack(fill=tk.X, pady=5)

        ttk.Label(time_country_frame, text="Pickup Date:").pack(side=tk.LEFT)
        today = datetime.datetime.now()
        dates = [(today + datetime.timedelta(days=i)).strftime("%Y-%m-%d") for i in range(14)]
        self.batch_date = ttk.Combobox(time_country_frame, values=dates, state="readonly", width=12)
        self.batch_date.set(dates[0])
        self.batch_date.pack(side=tk.LEFT, padx=5)
        ttk.Label(time_country_frame, text="Default Country:").pack(side=tk.LEFT)
        self.batch_country = ttk.Combobox(time_country_frame, values=["US", "CA"], width=5)
        self.batch_country.set("CA")
        self.batch_country.pack(side=tk.LEFT, padx=5)

        # Time selection
        time_frame = ttk.Frame(main_frame)
        time_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(time_frame, text="Ready Time:").pack(side=tk.LEFT)
        self.batch_ready_time = ttk.Combobox(time_frame, values=[f"{h:02d}00" for h in range(7, 21)], width=8)
        self.batch_ready_time.set("0900")
        self.batch_ready_time.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(time_frame, text="Close Time:").pack(side=tk.LEFT)
        self.batch_close_time = ttk.Combobox(time_frame, values=[f"{h:02d}00" for h in range(7, 23)], width=8)
        self.batch_close_time.set("1700")
        self.batch_close_time.pack(side=tk.LEFT, padx=5)

        # Progress Table
        ttk.Label(main_frame, text="Processing Progress:").pack(anchor=tk.W, pady=(10, 0))
        columns = ("Address", "Status", "PRN/Error")
        self.batch_tree = ttk.Treeview(main_frame, columns=columns, show="headings", height=8)
        self.batch_tree.heading("Address", text="Address")
        self.batch_tree.column("Address", width=300)
        self.batch_tree.heading("Status", text="Status")
        self.batch_tree.column("Status", width=100)
        self.batch_tree.heading("PRN/Error", text="PRN / Error Detail")
        self.batch_tree.column("PRN/Error", width=300)
        self.batch_tree.pack(fill=tk.BOTH, expand=True, pady=5)

        # Mode controls
        mode_frame = ttk.Frame(main_frame)
        mode_frame.pack(fill=tk.X, pady=5)
        ttk.Checkbutton(mode_frame, text="Manual Mode (Step-by-Step)", variable=self.manual_mode).pack(side=tk.LEFT)
        self.next_btn = ttk.Button(mode_frame, text="Next Item >>", command=lambda: self.next_step_event.set(), state=tk.DISABLED)
        self.next_btn.pack(side=tk.LEFT, padx=10)

        # Buttons Frame
        btn_frame = ttk.Frame(main_frame, padding=(0, 20))
        btn_frame.pack(fill=tk.X)
        
        ttk.Separator(btn_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(0, 20))

        self.run_batch_btn = ttk.Button(btn_frame, text="Process Batch", style="Accent.TButton", command=self.start_batch_thread)
        self.run_batch_btn.pack(side=tk.RIGHT)

        self.stop_batch_btn = ttk.Button(btn_frame, text="Stop", command=self.stop_current_batch, state=tk.DISABLED)
        self.stop_batch_btn.pack(side=tk.RIGHT, padx=10)

        history_btn = ttk.Button(btn_frame, text="View History & Cancel", command=self.show_history_window)
        history_btn.pack(side=tk.LEFT)

    def parse_and_fill(self):
        raw_address = self.address_text.get("1.0", tk.END).strip()
        if not raw_address:
            messagebox.showwarning("Warning", "Please paste an address first.")
            return
            
        parsed = parse_address_string(raw_address)
        if parsed:
            for key in ["Street", "Suite", "City", "State", "Zip", "Phone", "Country", "CompanyName", "ContactName"]:
                if key in parsed and parsed[key]:
                    # Map CompanyName and ContactName to the fields dict keys if they differ
                    # In setup_ui, they are CompanyName and ContactName
                    if key in self.fields:
                        self.fields[key].delete(0, tk.END)
                        self.fields[key].insert(0, parsed[key])
        else:
            messagebox.showerror("Error", "Could not parse address. Please enter manually.")

    def select_all_history(self, tree):
        tree.selection_set(tree.get_children())

    def get_province_offset(self, state):
        """Returns the hour offset from EST for Canadian provinces."""
        offsets = {
            "NL": 1.5,
            "NS": 1, "NB": 1, "PE": 1,
            "QC": 0, "ON": 0,
            "MB": -1, "SK": -1,
            "AB": -2, "NT": -2,
            "BC": -3, "YT": -3
        }
        return offsets.get(state.upper(), 0)

    def adjust_time_for_timezone(self, pickup_data):
        """
        Ensures the ReadyTime is not in the past for same-day pickups.
        UPS expects the 'Local Time' of the pickup address.
        """
        if pickup_data.get("PickupDate") != datetime.datetime.now().strftime("%Y%m%d"):
            return # Only matters for same-day
            
        state = pickup_data.get("State", "ON")
        offset = self.get_province_offset(state)
        
        # Current local time at pickup location
        now_est = datetime.datetime.now()
        now_local = now_est + datetime.timedelta(hours=offset)
        local_time_str = now_local.strftime("%H%M")
        
        requested_ready = pickup_data.get("ReadyTime", "0900")
        
        # If requested ready time is in the past (or within 30 mins), push it forward
        if int(requested_ready) < int(local_time_str) + 30:
            # Shift to 1 hour from now, rounded up to the next hour
            new_ready = (now_local + datetime.timedelta(hours=1)).strftime("%H00")
            print(f"[Timezone] Safety Push for {state} (Currently {local_time_str} Local) from {requested_ready} to {new_ready}")
            pickup_data["ReadyTime"] = new_ready
            
            # Ensure CloseTime is still at least 2 hours after ReadyTime
            close_time = pickup_data.get("CloseTime", "1700")
            if int(close_time) < int(new_ready) + 200:
                pickup_data["CloseTime"] = str(int(new_ready) + 200).zfill(4)
                if int(pickup_data["CloseTime"]) > 2200: pickup_data["CloseTime"] = "2200"

    def submit_pickup(self):
        pickup_data = {k: v.get() for k, v in self.fields.items()}
        # Reverted back to the requested prepaid tracking number
        pickup_data["TrackingNumber"] = "1Z4A059A9190837115"
        
        pickup_data["PickupDate"] = self.date_cb.get().replace("-", "")
        pickup_data["ReadyTime"] = self.ready_cb.get().replace(":", "")
        pickup_data["CloseTime"] = self.close_cb.get().replace(":", "")
        
        if messagebox.askyesno("Audit Confirmation", f"Schedule pickup for {pickup_data['Street']}?"):
            try:
                pickup_data["Country"] = pickup_data.get("Country", "").upper()
                pickup_data["State"] = pickup_data.get("State", "").upper()
                country = pickup_data["Country"]
                
                if country == "CA":
                    pickup_data["ServiceCode"] = "011"
                elif country == "US":
                    pickup_data["ServiceCode"] = "003"
                else:
                    pickup_data["ServiceCode"] = "007"

                # 1. Automate return label if no unique tracking is provided
                if pickup_data["TrackingNumber"] == "1Z4A059A9190837115":
                    # If it's the default, let's try to generate a unique one
                    label_res = self.api_client.create_return_label(pickup_data)
                    if label_res.get("status") == "success":
                        pickup_data["TrackingNumber"] = label_res["TrackingNumber"]
                    else:
                        print(f"[Warning] Automated label failed, using default: {label_res.get('message')}")

                # 2. Adjust for timezone if same-day
                self.adjust_time_for_timezone(pickup_data)

                # 3. Call the pickup API
                result = self.api_client.create_pickup(pickup_data)

                # 4. Extract PRN if success
                prn = ""
                if "PickupCreationResponse" in result and "PRN" in result["PickupCreationResponse"]:
                    prn = result["PickupCreationResponse"]["PRN"]

                # 5. Log to history IMMEDIATELY
                full_addr = f"{pickup_data.get('Street', '')}, {pickup_data.get('City', '')}, {pickup_data.get('State', '')} {pickup_data.get('Zip', '')}, {pickup_data.get('Country', '')}".strip(", ")
                entry = {
                    "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "company": pickup_data.get("CompanyName", ""),
                    "address": full_addr,
                    "date": pickup_data.get("PickupDate", ""),
                    "status": "Success" if prn else "Failed",
                    "prn": prn,
                    "details": pickup_data
                }
                self.save_to_history(entry)

                # 6. Now show UI feedback
                if prn:
                    self.status_var.set(f"Success: {prn}")
                    self.show_success_dialog(prn, pickup_data)
                elif "response" in result and "errors" in result.get("response", {}):
                    errors = result["response"]["errors"]
                    friendly_errors = []
                    for err in errors:
                        code = err.get("code")
                        msg = UPS_ERROR_MAP.get(code, err.get("message"))
                        friendly_errors.append(f"• [{code}] {msg}")
                    
                    error_info = "\n".join(friendly_errors)
                    self.status_var.set("Pickup Failed - Check Details")
                    messagebox.showerror("Pickup Rejection", f"The request was rejected by UPS:\n\n{error_info}")
                else:
                    msg = json.dumps(result, indent=2)
                    self.status_var.set("Unexpected Response")
                    messagebox.showinfo("API Response", f"Response details:\n{msg}")
            except Exception as e:
                logging.error(f"Exception in submit_pickup: {str(e)}")
                messagebox.showerror("API Error", str(e))

    def show_success_dialog(self, prn, pickup_data):
        top = tk.Toplevel(self.root)
        top.title("Success")
        top.geometry("420x400")
        top.configure(bg="white")
        top.transient(self.root)
        top.grab_set()

        header = ttk.Label(top, text="Pickup Scheduled", style="CardHeader.TLabel")
        header.pack(pady=(20, 10))

        msg = f"PRN: {prn}\n"
        msg += f"Date: {pickup_data.get('PickupDate')}\n"
        msg += f"Time: {pickup_data.get('ReadyTime')} - {pickup_data.get('CloseTime')}\n"
        msg += f"Address: {pickup_data.get('Street')}\n\n"
        msg += "Provide this to the client."
        
        txt = tk.Text(top, height=8, width=40, font=self.main_font, bg="#F5F5F7", borderwidth=0, padx=10, pady=10)
        txt.insert(tk.END, msg)
        txt.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)
        
        def copy_to_clipboard():
            self.root.clipboard_clear()
            self.root.clipboard_append(msg)
            self.status_var.set("Copied to Clipboard!")
            messagebox.showinfo("Copied", "Details copied.")

        btn_frame = ttk.Frame(top, padding=20)
        btn_frame.pack(fill=tk.X)

        copy_btn = ttk.Button(btn_frame, text="Copy Details", style="Accent.TButton", command=copy_to_clipboard)
        copy_btn.pack(side=tk.RIGHT)
        
        close_btn = ttk.Button(btn_frame, text="Close", command=top.destroy)
        close_btn.pack(side=tk.RIGHT, padx=10)

    def stop_current_batch(self):
        self.stop_batch = True
        
    def start_batch_thread(self):
        self.stop_batch = False
        raw_text = self.batch_text.get("1.0", tk.END).strip()
        if not raw_text:
            messagebox.showwarning("Warning", "Please paste addresses first.")
            return
            
        threading.Thread(target=self.process_batch, args=(raw_text,), daemon=True).start()

    def process_batch(self, raw_text):
        address_blocks = split_addresses(raw_text)
        if len(address_blocks) > 150:
            messagebox.showwarning("Limit Exceeded", f"Only first 150 addresses will be processed ({len(address_blocks)} found).")
            address_blocks = address_blocks[:150]

        # Clear previous progress
        for item in self.batch_tree.get_children():
            self.batch_tree.delete(item)

        defaults = {k: v.get() for k, v in self.batch_defaults.items()}
        common_date = self.batch_date.get().replace("-", "")
        default_country = self.batch_country.get()

        self.run_batch_btn.config(state=tk.DISABLED)
        self.stop_batch_btn.config(state=tk.NORMAL)
        
        ready_t = self.batch_ready_time.get()
        close_t = self.batch_close_time.get()
        
        for block in address_blocks:
            if self.stop_batch:
                self.log_message("Batch processing stopped by user.")
                break
                
            parsed = parse_address_string(block)
            if not parsed:
                self.batch_tree.insert("", tk.END, values=(block[:50]+"...", "Error", "Parsing Failed"))
                continue
                
            # Use batch default country if parser didn't find one
            if not parsed.get("Country"):
                 parsed["Country"] = default_country

            # Smart Service Code Selection based on Country
            service_code = defaults.get("ServiceCode", "011")
            country = parsed.get("Country", "")
            if country == "CA":
                service_code = "011" # UPS Standard for Canada
            elif country == "US":
                service_code = "003" # UPS Ground for US
            else:
                service_code = "007" # UPS Worldwide Express for international
            
            # Pre-validation — State not required for international addresses
            missing = []
            if not parsed.get("Street"): missing.append("Street")
            if not parsed.get("City"): missing.append("City")
            if not parsed.get("State") and country in ("CA", "US"): missing.append("State")
            if not parsed.get("Zip"): missing.append("Zip")
            
            if missing:
                self.batch_tree.insert("", tk.END, values=(block[:50]+"...", "Failed", f"Missing: {', '.join(missing)}"))
                continue

            # Merge with defaults
            pickup_data = {
                "Street": parsed.get("Street"),
                "Suite": parsed.get("Suite"),
                "City": parsed.get("City"),
                "State": parsed.get("State"),
                "Zip": parsed.get("Zip"),
                "Country": parsed.get("Country", "US"),
                "Phone": parsed.get("Phone") or defaults.get("Phone"),
                "CompanyName": defaults.get("CompanyName"),
                "ContactName": defaults.get("ContactName"),
                "ServiceCode": service_code,
                "Email": defaults.get("Email"),
                "PickupDate": common_date,
                "ReadyTime": ready_t,
                "CloseTime": close_t,
                "TrackingNumber": parsed.get("TrackingNumber") or "1Z4A059A9190837115"
            }

            iid = self.batch_tree.insert("", tk.END, values=(f"{pickup_data['Street']}, {pickup_data['City']}", "Processing...", ""))
            self.root.update_idletasks()

            try:
                # 1. Check if we need to generate a unique return label
                if not parsed.get("TrackingNumber"):
                    self.batch_tree.item(iid, values=(f"{pickup_data['Street']}, {pickup_data['City']}", "Creating Label...", ""))
                    label_res = self.api_client.create_return_label(pickup_data)
                    if label_res.get("status") == "success":
                        pickup_data["TrackingNumber"] = label_res["TrackingNumber"]
                    else:
                        err = label_res.get("message", "Label Error")
                        self.batch_tree.item(iid, values=(f"{pickup_data['Street']}, {pickup_data['City']}", "Failed (Label)", err))
                        continue

                # 2. Schedule the pickup
                self.adjust_time_for_timezone(pickup_data)
                self.batch_tree.item(iid, values=(f"{pickup_data['Street']}, {pickup_data['City']}", "Scheduling...", pickup_data["TrackingNumber"]))
                
                # Add a small delay for rate limiting
                time.sleep(1.5)
                
                result = self.api_client.create_pickup(pickup_data)
                prn = ""
                if "PickupCreationResponse" in result and "PRN" in result["PickupCreationResponse"]:
                    prn = result["PickupCreationResponse"]["PRN"]
                    self.log_message(f"Success for {pickup_data['Street']}: {prn}")
                    self.batch_tree.item(iid, values=(f"{pickup_data['Street']}, {pickup_data['City']}", "Success", prn))
                # Log to history
                full_addr = f"{pickup_data.get('Street', '')}, {pickup_data.get('City', '')}, {pickup_data.get('State', '')} {pickup_data.get('Zip', '')}, {pickup_data.get('Country', '')}".strip(", ")
                
                entry = {
                    "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "company": pickup_data["CompanyName"],
                    "address": full_addr,
                    "date": pickup_data["PickupDate"],
                    "status": "Success" if prn else "Failed",
                    "prn": prn,
                    "details": pickup_data
                }
                self.save_to_history(entry)
                if not prn:
                    logging.warning(f"Batch Failure for {pickup_data['Street']}: {err}")

            except Exception as e:
                logging.error(f"Exception in process_batch for {pickup_data.get('Street')}: {str(e)}")
                self.batch_tree.item(iid, values=(f"{pickup_data['Street']}, {pickup_data['City']}", "Error", str(e)))

        self.run_batch_btn.config(state=tk.NORMAL)
        self.stop_batch_btn.config(state=tk.DISABLED)
        self.next_btn.config(state=tk.DISABLED)
        messagebox.showinfo("Batch Complete", "Batch processing finished.")

    def cancel_pickup_gui(self):
        prn = simpledialog.askstring("Cancel Pickup", "Enter the PRN to cancel:", parent=self.root)
        if prn:
            try:
                result = self.api_client.cancel_pickup(prn)
                messagebox.showinfo("Cancel Result", json.dumps(result, indent=2))
            except Exception as e:
                messagebox.showerror("Cancel Error", str(e))

    def clear_fields(self):
        self.address_text.delete("1.0", tk.END)
        for entry in self.fields.values():
            entry.delete(0, tk.END)
        self.fields["Country"].insert(0, "CA")
        self.fields["CompanyName"].insert(0, "Omnitrans Inc")
        self.fields["ContactName"].insert(0, "Warehouse")

    def save_to_history(self, entry):
        history = []
        try:
            with open("pickup_history.json", "r") as f:
                history = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            history = []
        history.append(entry)
        with open(self.history_file, "w") as f:
            json.dump(history, f, indent=2)

    def show_history_window(self):
        top = tk.Toplevel(self.root)
        top.title("Pickup History")
        top.geometry("1000x500")
        
        # Search Frame
        search_frame = ttk.Frame(top, padding="10")
        search_frame.pack(fill=tk.X)
        ttk.Label(search_frame, text="Search (PRN/Address/Company):").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=40)
        search_entry.pack(side=tk.LEFT, padx=10)
        search_entry.focus_set()

        search_entry.pack(side=tk.LEFT, padx=10)
        search_entry.focus_set()

        history = []
        try:
            with open(self.history_file, "r") as f:
                history = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            pass
            
        columns = ("Timestamp", "Company", "Address", "Date", "Time", "Status", "PRN")
        tree_frame = ttk.Frame(top)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended")
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        def refresh_tree(*args):
            query = search_var.get().lower()
            for item in tree.get_children():
                tree.delete(item)
            for entry in history:
                match = any(query in str(entry.get(k, "")).lower() for k in ["prn", "address", "company"])
                if not query or match:
                    details = entry.get("details", {})
                    if isinstance(details, str):
                        try: details = json.loads(details)
                        except: details = {}
                    
                    ready_t = details.get("ReadyTime", "")
                    close_t = details.get("CloseTime", "")
                    time_str = f"{ready_t}-{close_t}" if ready_t else ""

                    tree.insert("", tk.END, values=(
                        entry.get("timestamp", ""),
                        entry.get("company", ""),
                        entry.get("address", ""),
                        entry.get("date", ""),
                        time_str,
                        entry.get("status", ""),
                        entry.get("prn", "")
                    ))

        search_var.trace_add("write", refresh_tree)
        
        for col in columns:
            tree.heading(col, text=col, command=lambda _col=col: self.sort_treeview(tree, _col, False))
            tree.column(col, width=120)
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Export Button
        export_btn = ttk.Button(search_frame, text="Export Selected to Excel", 
                               command=lambda: self.export_to_excel(tree, history))
        export_btn.pack(side=tk.LEFT, padx=10)
        
        def on_double_click(event):
            selection = tree.selection()
            if not selection:
                return
            item_iid = selection[0]
            values = tree.item(item_iid, "values")
            prn = values[6]  # Column 6 = PRN (0:Timestamp, 1:Company, 2:Address, 3:Date, 4:Time, 5:Status, 6:PRN)
            
            if prn:
                # Find the locally stored record first
                local_entry = next((e for e in history if e.get("prn") == prn), None)
                details = local_entry.get("details", {}) if local_entry else {}
                if isinstance(details, str):
                    try: details = json.loads(details)
                    except: details = {}

                # Build a detail window from local data
                details_win = tk.Toplevel(top)
                details_win.title(f"Pickup Details: {prn}")
                details_win.geometry("520x420")

                txt = tk.Text(details_win, padx=15, pady=10, font=self.main_font, bg="#F5F5F7", borderwidth=0)
                txt.pack(fill=tk.BOTH, expand=True)

                txt.insert(tk.END, f"PRN: {prn}\n")
                txt.insert(tk.END, f"Status: {local_entry.get('status', 'N/A') if local_entry else 'N/A'}\n")
                txt.insert(tk.END, f"Company: {local_entry.get('company', 'N/A') if local_entry else 'N/A'}\n")
                txt.insert(tk.END, f"Address: {local_entry.get('address', 'N/A') if local_entry else 'N/A'}\n")
                txt.insert(tk.END, f"Pickup Date: {details.get('PickupDate', local_entry.get('date', 'N/A') if local_entry else 'N/A')}\n")
                txt.insert(tk.END, f"Ready Time: {details.get('ReadyTime', 'N/A')}\n")
                txt.insert(tk.END, f"Close Time: {details.get('CloseTime', 'N/A')}\n")
                txt.insert(tk.END, f"Tracking #: {details.get('TrackingNumber', 'N/A')}\n")
                txt.insert(tk.END, f"Recorded At: {local_entry.get('timestamp', 'N/A') if local_entry else 'N/A'}\n")
                txt.insert(tk.END, "\n─────────────────────────────\n")
                txt.insert(tk.END, "Note: Live UPS status lookup requires additional API entitlements.\n")
                txt.insert(tk.END, "Showing locally recorded data.\n")
                txt.config(state=tk.DISABLED)
            else:
                # Fallback to Tracking search if no PRN
                entry = next((e for e in history if f"{e.get('Street', '')}, {e.get('City', '')}" == values[2]), None)
                if entry and entry.get("TrackingNumber"):
                    webbrowser.open(f"https://www.ups.com/track?tracknum={entry['TrackingNumber']}")

        tree.bind("<Double-1>", on_double_click)
        
        # UI Hint for Range Selection
        ttk.Label(top, text="Tip: Use Shift+Click to select a range of items, or Ctrl+Click to pick specific ones.", 
                  font=("Helvetica", 9, "italic"), foreground="gray").pack(side=tk.BOTTOM, pady=5)

        # Initial load
        refresh_tree()

    def sort_treeview(self, tree, col, reverse):
        l = [(tree.set(k, col), k) for k in tree.get_children('')]
        l.sort(reverse=reverse)
        for index, (val, k) in enumerate(l):
            tree.move(k, '', index)
        tree.heading(col, command=lambda: self.sort_treeview(tree, col, not reverse))

    def export_to_excel(self, tree, history):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Export", "Please select one or more items to export.")
            return
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UPS Pickups"
        
        headers = ["Created", "Company", "Full Address", "Date", "Ready", "Close", "Status", "PRN", "Tracking #"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            
        for row_idx, item_iid in enumerate(selected, 2):
            values = tree.item(item_iid, "values")
            # values: (Timestamp, Company, Address, Date, Status, PRN)
            
            # More robust lookup: Match by timestamp and address
            entry = next((e for e in history if e.get("timestamp") == values[0] and e.get("address") == values[2]), None)
            
            if entry:
                details = entry.get("details", {})
                # Handle cases where details might be a JSON string or not a dict
                if isinstance(details, str):
                    try:
                        details = json.loads(details)
                    except:
                        details = {}
                if not isinstance(details, dict):
                    details = {}
                
                # Full Address: Rebuild with all components for completeness
                street = details.get("Street", "")
                city = details.get("City", "")
                state = details.get("State", "")
                zip_code = details.get("Zip", "")
                country = details.get("Country", "")
                
                if street:
                    full_address = f"{street}, {city}, {state} {zip_code}, {country}".strip(", ")
                else:
                    # Fallback to the top-level address if details are missing
                    full_address = entry.get("address", "")

                row_data = [
                    entry.get("timestamp"),
                    entry.get("company"),
                    full_address,
                    entry.get("date"),
                    details.get("ReadyTime"),
                    details.get("CloseTime"),
                    entry.get("status"),
                    entry.get("prn"),
                    details.get("TrackingNumber")
                ]
            else:
                # Fallback: Use values directly from Treeview if history match fails
                row_data = [
                    values[0], # Timestamp
                    values[1], # Company
                    values[2], # Full Address from Treeview
                    values[3], # Date
                    "", "", # Times placeholder
                    values[4], # Status
                    values[5], # PRN
                    "" # Tracking placeholder
                ]
                
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
        
        filename = f"UPS_Pickups_Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            wb.save(filename)
            messagebox.showinfo("Export Success", f"Exported {len(selected)} items to {filename}")
        except PermissionError:
            messagebox.showerror("Export Error", f"Could not save {filename}. Please close the file if it is open in Excel and try again.")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save file: {str(e)}")
            
        def cancel_selected():
            selected_items = tree.selection()
            if not selected_items: 
                messagebox.showwarning("Warning", "Please select at least one pickup to cancel.")
                return
            
            if not messagebox.askyesno("Confirm", f"Cancel {len(selected_items)} selected pickup(s)?"):
                return

            success_count = 0
            fail_count = 0
            
            for item_iid in selected_items:
                values = tree.item(item_iid, "values")
                prn = values[6]  # PRN is index 6 (0:Timestamp,1:Company,2:Address,3:Date,4:Time,5:Status,6:PRN)
                
                if prn:
                    # Find the entry in history
                    entry = next((e for e in history if e.get("prn") == prn), None)
                    if entry:
                        try:
                            res = self.api_client.cancel_pickup(prn)
                            if res.get("status") == "success" or "error" not in str(res).lower():
                                success_count += 1
                                tree.set(item_iid, "Status", "Cancelled")
                                entry["status"] = "Cancelled"
                            else:
                                fail_count += 1
                        except Exception as e:
                            fail_count += 1
                    else:
                        fail_count += 1
                else:
                    fail_count += 1

            # Update history file
            with open("pickup_history.json", "w") as f:
                json.dump(history, f, indent=2)

            messagebox.showinfo("Bulk Cancel Result", f"Finished.\nSuccess: {success_count}\nFailed: {fail_count}")

        def select_all():
            tree.selection_set(tree.get_children())

        def cancel_all_btn():
            select_all()
            cancel_selected()

        def clear_failed_cancelled():
            nonlocal history
            if not messagebox.askyesno("Confirm", "Remove all records with Status 'Failed' or 'Cancelled'?"):
                return
            new_history = [e for e in history if e.get("status") not in ["Failed", "Cancelled"]]
            deleted_count = len(history) - len(new_history)
            history[:] = new_history # Update in place
            with open(self.history_file, "w") as f:
                json.dump(history, f, indent=2)
            refresh_tree()
            messagebox.showinfo("Success", f"Cleared {deleted_count} record(s).")

        btn_frame = ttk.Frame(top)
        btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text="Select All", command=select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel All", command=cancel_all_btn).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel Selected Pickup(s)", command=cancel_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Clear Failed/Cancelled", command=clear_failed_cancelled).pack(side=tk.LEFT, padx=5)

if __name__ == "__main__":
    root = tk.Tk()
    app = UPSPickupGUI(root)
    root.mainloop()
